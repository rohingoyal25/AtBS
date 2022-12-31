import openpyxl
from openpyxl.utils import get_column_letter
import sys
import time

def main():
    if len(sys.argv) < 3:
        sys.exit("Incorrect number of command line arguments. Put input file first and output file second.")

    wb_inp_name = sys.argv[1]
    wb_out_name = sys.argv[2]

    wb_inp = openpyxl.load_workbook(wb_inp_name)
    wb_out = openpyxl.Workbook()
    sheet_inp = wb_inp.active
    sheet_out = wb_out.active
    # sheet_data = []
    # for i in range(1, sheet_inp.max_column + 1):
    #     tmp = []
    #     for j in range(1, sheet_inp.max_row + 1):
    #         tmp.append(sheet_inp[get_column_letter(i) + str(j)].value)
    #     sheet_data.append(tmp)

    print(f"Max row: {sheet_inp.max_row}")
    print('Working...')
    tic = time.perf_counter()
    for row_num in range(1, 18279):
        for col_num in range(1, sheet_inp.max_column + 1):
            col = get_column_letter(col_num)
            row = get_column_letter(row_num)
            sheet_out[row + str(col_num)] = sheet_inp[col + str(row_num)].value
        print(f'{time.perf_counter() - tic}')

    wb_out.save(wb_out_name)
    # print(sheet_data)

main()