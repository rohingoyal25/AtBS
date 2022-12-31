import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

def main():
    wb = openpyxl.Workbook()
    sheet = wb.active
    font_obj = Font(bold = True)
    for i in range(2, 8):
        col_let = get_column_letter(i)

        sheet['A' + str(i)] = i-1
        sheet['A' + str(i)].font = font_obj

        sheet[col_let + str(1)] = i - 1
        sheet[col_let + str(1)].font = font_obj

    sheet.freeze_panes = 'A2'
    for row_num in range(2, 8):
        for col_num in range(2, 8):
            sheet[get_column_letter(col_num) + str(row_num)] = (row_num - 1) * (col_num - 1)
    wb.save('multable.xlsx')

main()