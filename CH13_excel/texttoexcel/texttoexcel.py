import openpyxl
from openpyxl.utils import get_column_letter
import sys

def main():
    if len(sys.argv) < 3:
        sys.exit('Need at least 1 txt file to put into spreadsheet and 1 xlsx file to output to')
    elif not sys.argv[-1].endswith('.xlsx'):
        sys.exit('Final argument must be name of an excel file to output to')
    else:
        for item in sys.argv[1:-1]:
            if not item.endswith('.txt'):
                sys.exit('Files to read dont end in .txt')

    wb = openpyxl.Workbook()
    sheet = wb.active

    for i, txt in enumerate(sys.argv[1:-1], 1):
        col = get_column_letter(i)
        with open(txt) as file:
            lines = file.readlines()
            for j, line in enumerate(lines, 1):
                sheet[col + str(j)] = line
    wb.save(sys.argv[-1])

main()